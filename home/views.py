from openpyxl import load_workbook
from django.shortcuts import render

# Create your views here.


def index(request):
    """ A view to navigate to the index page """

    workbook = load_workbook('static/excel-files/EnergyConsumptionDetail_updated.xlsx')
    sheet_names = workbook.sheetnames[0:4]

    """ Rate detail """
    example_sheet = workbook["Example"]
    rate1 = example_sheet['G6'].value
    rate2 = example_sheet['G7'].value
    rate3 = example_sheet['G8'].value

    """ Customer 1 """
    sheet1 = workbook["Customer 1"]
    sheet_name1 = workbook.sheetnames[0]
    for value in sheet1.iter_cols(min_row=1,
                                  max_row=7,
                                  min_col=1,
                                  max_col=1,
                                  values_only=True):
        headings1 = value
    
    for value in sheet1.iter_cols(min_row=1,
                                  max_row=7,
                                  min_col=2,
                                  max_col=2,
                                  values_only=True):
        first_row1 = value
    
    for value in sheet1.iter_cols(min_row=1,
                                  max_row=7,
                                  min_col=3,
                                  max_col=3,
                                  values_only=True):
        second_row1 = value
    
    """ Total energy cost for customer 1"""

    c1_day_rate1 = sheet1['B6'].value
    c1_day_rate2 = sheet1['C6'].value
    c1_night_rate1 = sheet1['B7'].value
    c1_night_rate2 = sheet1['C7'].value
    c1_energyconsupt1 = abs(c1_day_rate1 - c1_day_rate2)
    c1_energyconsupt2 = abs(c1_night_rate1 - c1_night_rate2)
    c1_rate1_consupt = round(c1_energyconsupt1 * rate1)
    c1_rate2_consupt = round(c1_energyconsupt2 * rate2)
    total_energy_cost1 = c1_rate1_consupt + c1_rate2_consupt

    """ Customer 2 """
    sheet2 = workbook["Customer 2"]
    sheet_name2 = workbook.sheetnames[1]
    for value in sheet2.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=1,
                                  max_col=1,
                                  values_only=True):
        headings2 = value
    
    for value in sheet2.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=2,
                                  max_col=2,
                                  values_only=True):
        first_row2 = value
    
    for value in sheet2.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=3,
                                  max_col=3,
                                  values_only=True):
        second_row2 = value
    
    """ Total energy cost for customer 2"""

    c2_day_rate1 = sheet2['B6'].value
    c2_day_rate2 = sheet2['C6'].value
    c2_night_rate1 = sheet2['B7'].value
    c2_night_rate2 = sheet2['C7'].value
    c2_weekend_rate1 = sheet2['B8'].value
    c2_weekend_rate2 = sheet2['C8'].value
    c2_energyconsupt1 = abs(c2_day_rate1 - c2_day_rate2)
    c2_energyconsupt2 = abs(c2_night_rate1 - c2_night_rate2)
    c2_energyconsupt3 = abs(c2_weekend_rate1 - c2_weekend_rate2)
    c2_rate1_consupt = round(c2_energyconsupt1 * rate1)
    c2_rate2_consupt = round(c2_energyconsupt2 * rate2)
    c2_rate3_consupt = round(c2_energyconsupt3 * rate3)
    total_energy_cost2 = c2_rate1_consupt + c2_rate2_consupt + c2_rate3_consupt

    """ Customer 3 """
    sheet3 = workbook["Customer 3"]
    sheet_name3 = workbook.sheetnames[2]
    for value in sheet3.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=1,
                                  max_col=1,
                                  values_only=True):
        headings3 = value
    
    for value in sheet3.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=2,
                                  max_col=2,
                                  values_only=True):
        first_row3 = value
    
    for value in sheet3.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=3,
                                  max_col=3,
                                  values_only=True):
        second_row3 = value

    """ Total energy cost for customer 3"""

    c3_day_rate1 = sheet3['B6'].value
    c3_day_rate2 = sheet3['C6'].value
    c3_night_rate1 = sheet3['B7'].value
    c3_night_rate2 = sheet3['C7'].value
    c3_weekend_rate1 = sheet3['B8'].value
    c3_weekend_rate2 = sheet3['C8'].value
    c3_energyconsupt1 = abs(c3_day_rate1 - c3_day_rate2)
    c3_energyconsupt2 = abs(c3_night_rate1 - c3_night_rate2)
    c3_energyconsupt3 = abs(c3_weekend_rate1 - c3_weekend_rate2)
    c3_rate1_consupt = round(c3_energyconsupt1 * rate1)
    c3_rate2_consupt = round(c3_energyconsupt2 * rate2)
    c3_rate3_consupt = round(c3_energyconsupt3 * rate3)
    total_energy_cost3 = c3_rate1_consupt + c3_rate2_consupt + c3_rate3_consupt

    """ Customer 4 """
    sheet4 = workbook["Customer 4"]
    sheet_name4 = workbook.sheetnames[3]
    for value in sheet4.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=1,
                                  max_col=1,
                                  values_only=True):
        headings4 = value
    
    for value in sheet4.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=2,
                                  max_col=2,
                                  values_only=True):
        first_row4 = value
    
    for value in sheet4.iter_cols(min_row=1,
                                  max_row=8,
                                  min_col=3,
                                  max_col=3,
                                  values_only=True):
        second_row4 = value

    """ Total energy cost for customer 4"""

    c4_day_rate1 = sheet4['B6'].value
    c4_day_rate2 = sheet4['C6'].value
    c4_weekend_rate1 = sheet4['B7'].value
    c4_weekend_rate2 = sheet4['C7'].value
    c4_energyconsupt1 = abs(c4_day_rate1 - c4_day_rate2)
    c4_energyconsupt2 = abs(c4_weekend_rate1 - c4_weekend_rate2)
    c4_rate1_consupt = round(c4_energyconsupt1 * rate1)
    c4_rate2_consupt = round(c4_energyconsupt2 * rate2)
    total_energy_cost4 = c4_rate1_consupt + c4_rate2_consupt

    """ Highest day rate consumption """

    if c1_energyconsupt1 > max(c2_energyconsupt1, c3_energyconsupt1, c4_energyconsupt1):
        highest_day_customer = sheet_name1
        highest_day_headings = headings1
        highest_day_row1 = first_row1
        highest_day_row2 = second_row1
        highest_day_energy = total_energy_cost1
    elif c2_energyconsupt1 > max(c1_energyconsupt1, c3_energyconsupt1, c4_energyconsupt1):
        highest_day_customer = sheet_name2
        highest_day_headings = headings2
        highest_day_row1 = first_row2
        highest_day_row2 = second_row2
        highest_day_energy = total_energy_cost2
    elif c3_energyconsupt1 > max(c1_energyconsupt1, c2_energyconsupt1,
                                 c4_energyconsupt1):
        highest_day_customer = sheet_name3
        highest_day_headings = headings3
        highest_day_row1 = first_row3
        highest_day_row2 = second_row3
        highest_day_energy = total_energy_cost3
    elif c4_energyconsupt1 > max(c1_energyconsupt1, c2_energyconsupt1,
                                 c3_energyconsupt1):
        highest_day_customer = sheet_name4
        highest_day_headings = headings4
        highest_day_row1 = first_row4
        highest_day_row2 = second_row4
        highest_day_energy = total_energy_cost4

    """ Highest night rate consumption """

    if c1_energyconsupt2 > max(c2_energyconsupt2, c3_energyconsupt2,
                               c4_energyconsupt2):
        highest_night_customer = sheet_name1
        highest_night_headings = headings1
        highest_night_row1 = first_row1
        highest_night_row2 = second_row1
        highest_night_energy = total_energy_cost1
    elif c2_energyconsupt2 > max(c1_energyconsupt2, c3_energyconsupt2,
                                 c4_energyconsupt2):
        highest_night_customer = sheet_name2
        highest_night_headings = headings2
        highest_night_row1 = first_row2
        highest_night_row2 = second_row2
        highest_night_energy = total_energy_cost2
    elif c3_energyconsupt2 > max(c1_energyconsupt2, c2_energyconsupt2,
                                 c4_energyconsupt1):
        highest_night_customer = sheet_name3
        highest_night_headings = headings3
        highest_night_row1 = first_row3
        highest_night_row2 = second_row3
        highest_night_energy = total_energy_cost3
    elif c4_energyconsupt2 > max(c1_energyconsupt2, c2_energyconsupt2,
                                 c3_energyconsupt2):
        highest_night_customer = sheet_name4
        highest_night_headings = headings4
        highest_night_row1 = first_row4
        highest_night_row2 = second_row4
        highest_night_energy = total_energy_cost4

    """ Highest total energy cost """

    if total_energy_cost1 > max(total_energy_cost2, total_energy_cost3,
                                total_energy_cost4):
        highest_energy_customer = sheet_name1
        highest_energy_headings = headings1
        highest_energy_row1 = first_row1
        highest_energy_row2 = second_row1
        highest_energy = total_energy_cost1
    elif total_energy_cost2 > max(total_energy_cost1, total_energy_cost3,
                                  total_energy_cost4):
        highest_energy_customer = sheet_name2
        highest_energy_headings = headings2
        highest_energy_row1 = first_row2
        highest_energy_row2 = second_row2
        highest_energy = total_energy_cost2
    elif total_energy_cost3 > max(total_energy_cost1, total_energy_cost2,
                                  total_energy_cost4):
        highest_energy_customer = sheet_name3
        highest_energy_headings = headings3
        highest_energy_row1 = first_row3
        highest_energy_row2 = second_row3
        highest_energy = total_energy_cost3
    elif total_energy_cost4 > max(total_energy_cost1, total_energy_cost2,
                                  total_energy_cost3):
        highest_energy_customer = sheet_name4
        highest_energy_headings = headings4
        highest_energy_row1 = first_row4
        highest_energy_row2 = second_row4
        highest_energy = total_energy_cost4

    context = {
        'sheet_names': sheet_names,
        'sheet_name1': sheet_name1,
        'headings1': headings1,
        'first_row1': first_row1,
        'second_row1': second_row1,
        'sheet_name2': sheet_name2,
        'headings2': headings2,
        'first_row2': first_row2,
        'second_row2': second_row2,
        'sheet_name3': sheet_name3,
        'headings3': headings3,
        'first_row3': first_row3,
        'second_row3': second_row3,
        'sheet_name4': sheet_name4,
        'headings4': headings4,
        'first_row4': first_row4,
        'second_row4': second_row4,
        'total_energy_cost1': total_energy_cost1,
        'total_energy_cost2': total_energy_cost2,
        'total_energy_cost3': total_energy_cost3,
        'total_energy_cost4': total_energy_cost4,
        'highest_day_customer': highest_day_customer,
        'highest_day_headings': highest_day_headings,
        'highest_day_row1': highest_day_row1,
        'highest_day_row2': highest_day_row2,
        'highest_day_energy': highest_day_energy,
        'highest_night_customer': highest_night_customer,
        'highest_night_headings': highest_night_headings,
        'highest_night_row1': highest_night_row1,
        'highest_night_row2': highest_night_row2,
        'highest_night_energy': highest_night_energy,
        'highest_energy_customer': highest_energy_customer,
        'highest_energy_headings': highest_energy_headings,
        'highest_energy_row1': highest_energy_row1,
        'highest_energy_row2': highest_energy_row2,
        'highest_energy': highest_energy,
    }

    return render(request, 'home/index.html', context)
